import openpyxl

def count_matching_values(file_path):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Initialize count
    count = 0

    # Iterate through each row
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=8, max_col=8):
        # Lowercase the values in columns H and P before comparing
        value_h = str(row[0].value).lower()
        value_p = str(sheet.cell(row=row[0].row, column=16).value).lower()

        # Check if the lowercase value in column H matches the lowercase value in column P
        if value_h == value_p:
            count += 1

    return count

# Example usage
file_path = 'detection_cccd_v3.1_20k.xlsx'
matching_count = count_matching_values(file_path)
print("Number of rows where column H and column P have same values (case-insensitive):", matching_count)


